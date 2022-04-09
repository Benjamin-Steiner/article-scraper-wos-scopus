import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import random
import re

# open excel
workbook = openpyxl.load_workbook("file_with_wos_articles.xlsx")
options = Options()

# Path to ChromeDriver
path= "C:\Program Files (x86)\chromedriver.exe"
driver=webdriver.Chrome(path,options=options)

# institution login URL
url=""

driver.get(url)

# institution Username and password
username=""
password=""

xpath="/html/body/div/table[1]/tbody/tr/td/form/table/tbody/tr/td/div/table/tbody/tr[1]/td[2]/input"

# insert username
username_field= driver.find_element_by_name('user')
username_field.send_keys(username)

# insert password
password_field= driver.find_element_by_name('pass')
password_field.send_keys(password)
time.sleep(random.randint(1, 5))

# Confirm
password_field.send_keys(Keys.RETURN)

time.sleep(random.randint(5, 10))

# accept cookies
cookie=driver.find_element_by_id('onetrust-accept-btn-handler').click()

# find search bar
search = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.NAME, 'search-main-box'))
    )

# press got it
box=driver.find_element_by_id('pendo-button-e5808a4c').click()

# Create a list of names of all worksheets in `workbook`
sheet_names = workbook.sheetnames

sheet_obj = workbook.active
m_row = sheet_obj.max_row

once=0
for i in range(2, m_row + 1):
    print (f"{i-1}")
    naslov_str = ""
    result = []
    title = ""
    new_str=""
    title = sheet_obj.cell(row=i, column=6)

    # Wos does not like special characters - so we just replaced them with spaces
    new_str = re.sub(r'[^A-Za-z0-9 ]+', ' ', title.value)
    new_str = re.sub(' +', ' ', new_str)
    result.append(new_str)
    try:
        if i==2:
            search.send_keys(new_str)
            search.send_keys(Keys.RETURN)
        else:
            # clear string before starting new search
            clear = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, 'div.button-row'))
            )
            try:
                clear_click = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//button[@data-ta="clear-search"]'))
                )
                clear_click.click()
            except:
                ...
            insert_title_new = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, '//input[@data-ta="search-criteria-input"]'))
            )
            # insert search string
            insert_title_new.send_keys(new_str)
            time.sleep(random.randint(1, 3))
            insert_title_new.send_keys(Keys.RETURN)

        # find first article
        prvi_clanek=None
        prvi_clanek = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".summary-record"))
        )

        time.sleep(1)

        try:
            if (once < 2):
                sign_in_btn = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "button.bb-button._pendo-button-secondaryButton._pendo-button"))
                )
                sign_in_btn.click()
                once+=1
        except:
            ...
        # add first article to the list
        prvi_clanek=prvi_clanek=driver.find_element_by_css_selector('.summary-record')

        check_box = prvi_clanek.find_element_by_class_name('mat-checkbox-inner-container').click()

        add_to_list = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#snRecListTop > app-marked-list-option > button"))
        )
        add_to_list.click()
        naslov = WebDriverWait(prvi_clanek, 20).until(
            EC.presence_of_element_located((By.XPATH, '//a[@data-ta="summary-record-title-link"]'))
        )
        link_url=naslov.get_attribute('href')
        naslov=naslov.get_attribute('text')

        result.append(naslov)
        result.append(link_url)

        # go to search box again
        box_search_new = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'button.search-terms'))
        )
        box_search_new.click()


        try:
            if (once<2):
                sign_in_btn = WebDriverWait(driver, 6).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "button.bb-button._pendo-button-secondaryButton._pendo-button"))
                )
                sign_in_btn.click()
                once += 1
        except:
            ...

        clear = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'div.button-row'))
        )
        clear_click = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, '//button[@data-ta="clear-search"]'))
        )
        clear_click.click()

        with open('added_documents.txt', 'a', encoding='utf-8') as f:
            for line in result:
                f.write(line)
                f.write('---')
            f.write('\n')
    except:
        title=""
        title = sheet_obj.cell(row=i, column=6)
        print('napaka pri line',i)
        with open('added_documents.txt', 'a', encoding='utf-8') as f:
            f.write(f'error at line: {i-1}')
            f.write('\n')
        driver.get('https://www.webofscience.com/wos/woscc/basic-search')

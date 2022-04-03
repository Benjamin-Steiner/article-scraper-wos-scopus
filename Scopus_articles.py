import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import random
from selenium.webdriver.common.action_chains import ActionChains

#open excel
workbook = openpyxl.load_workbook("file_with_scopus_articles.xlsx")
options = Options()

#Path to ChromeDriver
path= "C:\Program Files (x86)\chromedriver.exe"
driver=webdriver.Chrome(path,options=options)

#institution login URL
url=""

driver.get(url)

# institution Username and password
username=""
password=""

xpath="/html/body/div/table[1]/tbody/tr/td/form/table/tbody/tr/td/div/table/tbody/tr[1]/td[2]/input"


#insert username
username_field= driver.find_element_by_name('user')
username_field.send_keys(username)

#insert password
password_field= driver.find_element_by_name('pass')
password_field.send_keys(password)
time.sleep(random.randint(1, 5))

#Confirm
password_field.send_keys(Keys.RETURN)

time.sleep(random.randint(5, 10))




# Create a list of names of all worksheets in `workbook`
sheet_names = workbook.sheetnames

sheet_obj = workbook.active
m_row = sheet_obj.max_row

once=0
for i in range(2, m_row + 1):
    print (f"{i-1}")
    naslov_str = ""
    result = []
    title = ""
    new_str=""
    title = sheet_obj.cell(row=i, column=6)
    new_str = title.value

    # Search for exact string
    search_string = f'TITLE ( "{new_str}" )'
    result.append(new_str)
    try:
        driver.get('https://www.scopus.com/search/form.uri?zone=TopNavBar&origin=resultslist&display=advanced')

        # if there is a clear btn on the page, it will click it
        clear_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'clearLink'))
        )
        try:
            clear_field.click()
        except:
            ...

        # check if there is a search box on the page
        search_field = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, 'SCAdvSearchInputBox'))
        )
        search_field.click()

        # pinsert string
        actions = ActionChains(driver)
        actions.send_keys(search_string)
        actions.perform()

        time.sleep(random.randint(1, 3))

        # press enter
        actions = ActionChains(driver)
        actions.send_keys(Keys.RETURN)
        actions.perform()

        time.sleep(random.randint(1, 3))

        #time.sleep(1)
        prvi_clanek=None



        pravi_clanek = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.LINK_TEXT, new_str))
        )
        parent_of_url=pravi_clanek.find_element_by_xpath('..')
        correct_data_row=parent_of_url.find_element_by_xpath('..')
        correct_data_row.find_element_by_css_selector('.checkbox').click()

        url_clanka =pravi_clanek.get_attribute('href')
        naslov=pravi_clanek.get_attribute('text')

        result.append(naslov)
        result.append(url_clanka)

        add_to_list_btn = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "addToMyListResults"))
        )
        add_to_list_btn.click()
        time.sleep(random.randint(1, 3))
        # open txt file and save titles and URL of articles
        with open('added_documents.txt', 'a', encoding='utf-8') as f:
            for line in result:
                f.write(line)
                f.write('---')
            f.write('\n')
    except:
        title = ""
        title = sheet_obj.cell(row=i, column=6)
        print('Error at line ', i)
        with open('added_documents.txt', 'a', encoding='utf-8') as f:
            f.write(f'error at line: {i - 1}')
            f.write('\n')
        driver.get('https://www.scopus.com/search/form.uri?zone=TopNavBar&origin=resultslist&display=advanced')
